"""Vues principales de l'application ``core``.

Ces vues fournissent :
  - une page d'accueil marketing utilisant le thème existant ;
  - une page "à propos" statique ;
  - un endpoint de santé JSON pour le monitoring ;
  - un tableau de bord agrégé pour le back‑office.
"""

from datetime import timedelta
from io import BytesIO
from pathlib import Path

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template.response import TemplateResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.contrib.auth.models import User
from django.db.models import Sum

from services.models import Service, Category
from devis.models import Quote, Client
from factures.models import Invoice
from tasks.models import Task
from messaging.models import EmailMessage
from .services.document_service import ClientDocumentService
from .decorators import client_portal_required, worker_portal_required, admin_portal_required


def _get_client_portal_user(client):
    """Return the portal account matching a client email, if any."""
    return User.objects.filter(email__iexact=client.email).select_related('profile').first()


def _get_client_portal_record(user):
    """Return the latest client record linked to the authenticated portal user."""
    user_email = (getattr(user, "email", "") or "").strip()
    if not user_email:
        return None
    return Client.objects.filter(email__iexact=user_email).order_by("-created_at").first()


def _sum_amount(queryset, field_name):
    """Aggregate a monetary field and always return a concrete value."""
    return queryset.aggregate(total=Sum(field_name)).get("total") or 0


def _build_quote_summary(quotes):
    """Compute premium summary data for client quote pages."""
    today = timezone.localdate()
    pending_statuses = [Quote.QuoteStatus.DRAFT, Quote.QuoteStatus.SENT]
    return {
        "total": quotes.count(),
        "pending": quotes.filter(status__in=pending_statuses).count(),
        "accepted": quotes.filter(status=Quote.QuoteStatus.ACCEPTED).count(),
        "rejected": quotes.filter(status=Quote.QuoteStatus.REJECTED).count(),
        "expiring_soon": quotes.filter(
            status__in=pending_statuses,
            valid_until__isnull=False,
            valid_until__gte=today,
            valid_until__lte=today + timedelta(days=10),
        ).count(),
        "total_value": _sum_amount(quotes, "total_ttc"),
        "latest_issue_date": quotes.order_by("-issue_date").values_list("issue_date", flat=True).first(),
    }


def _build_invoice_summary(invoices):
    """Compute premium summary data for client invoice pages."""
    unpaid_statuses = [
        Invoice.InvoiceStatus.SENT,
        Invoice.InvoiceStatus.PARTIAL,
        Invoice.InvoiceStatus.OVERDUE,
    ]
    unpaid_invoices = invoices.filter(status__in=unpaid_statuses)
    return {
        "total": invoices.count(),
        "unpaid": unpaid_invoices.count(),
        "paid": invoices.filter(status=Invoice.InvoiceStatus.PAID).count(),
        "overdue": invoices.filter(status=Invoice.InvoiceStatus.OVERDUE).count(),
        "total_due": _sum_amount(unpaid_invoices, "total_ttc"),
        "total_paid": _sum_amount(invoices.filter(status=Invoice.InvoiceStatus.PAID), "total_ttc"),
        "next_due_date": unpaid_invoices.filter(due_date__isnull=False).order_by("due_date").values_list("due_date", flat=True).first(),
    }


def _build_document_summary(documents):
    """Compute premium summary data for manually published client documents."""
    today = timezone.localdate()
    return {
        "total": documents.count(),
        "pinned": documents.filter(is_pinned=True).count(),
        "expiring_soon": documents.filter(
            expires_at__isnull=False,
            expires_at__gte=today,
            expires_at__lte=today + timedelta(days=14),
        ).count(),
        "downloaded": documents.filter(download_count__gt=0).count(),
        "latest_publication": documents.order_by("-published_at").values_list("published_at", flat=True).first(),
    }


def home(request):
    """Page d'accueil publique."""
    from .models import Realisation

    categories = Category.objects.all().order_by("name")
    featured_services = (
        Service.objects.filter(is_active=True)
        .order_by("-created_at")[:6]
    )
    featured_realisations = Realisation.objects.filter(is_published=True)[:4]
    return render(
        request,
        "core/home.html",
        {
            "categories": categories,
            "featured_services": featured_services,
            "featured_realisations": featured_realisations,
        },
    )


def about(request):
    """Page de présentation de l'entreprise."""
    return render(request, "core/about.html")


def excellence(request):
    """
    Page de mise en avant des engagements de l'entreprise.

    Cette vue ne nécessite pas de logique complexe ; elle se contente
    d'afficher un template statique décrivant les valeurs et la
    philosophie de NetExpress.
    """
    return render(request, "core/excellence.html")


def legal_notice(request):
    """Mentions légales (obligation LCEN art. 6-III)."""
    branding = getattr(settings, "INVOICE_BRANDING", {})
    # Le numéro RCS d'une société correspond à son SIREN, soit les 9 premiers
    # chiffres du SIRET.
    siret_digits = "".join(c for c in str(branding.get("siret", "")) if c.isdigit())
    siren = siret_digits[:9] if len(siret_digits) >= 9 else ""
    return render(request, "core/legal_notice.html", {"branding": branding, "siren": siren})


def privacy_policy(request):
    """Politique de confidentialité (information des personnes, RGPD art. 13-14)."""
    return render(request, "core/privacy_policy.html", {"branding": getattr(settings, "INVOICE_BRANDING", {})})


def realisations(request):
    """Galerie des réalisations, alimentée depuis la base de données."""
    from .models import Realisation

    images = Realisation.objects.filter(is_published=True)

    # Construire la liste des catégories dynamiquement
    used_categories = set(images.values_list('category', flat=True))
    all_categories = [
        ('nettoyage', 'Nettoyage'),
        ('espaces_verts', 'Espaces verts'),
        ('renovation', 'Rénovation'),
        ('autre', 'Autre'),
    ]
    categories = [("all", "Toutes")] + [(slug, name) for slug, name in all_categories if slug in used_categories]

    return render(
        request,
        "core/realisations.html",
        {
            "images": images,
            "categories": categories,
        },
    )


def health(request):
    """Endpoint simple pour les probes de monitoring."""
    return JsonResponse({"status": "ok"})


# Dashboard technique supprimé - Migration vers /admin-dashboard/
# Les fonctionnalités sont maintenant disponibles dans admin_dashboard()
# Le portail métier principal reste /admin-dashboard/ ; /gestion/ ne sert qu'aux opérations techniques avancées.
# Les admins business accèdent à /admin-dashboard/


@client_portal_required
def client_dashboard(request):
    """Enhanced Client Portal dashboard with document filtering."""
    from .services.client_service import ClientService

    # Update last portal access
    if hasattr(request.user, 'profile'):
        from django.utils import timezone
        request.user.profile.last_portal_access = timezone.now()
        request.user.profile.save(update_fields=['last_portal_access'])
    
    client_record = _get_client_portal_record(request.user)

    # Get document statistics and recent documents using the service
    stats = ClientDocumentService.get_client_document_stats(request.user)
    recent_docs = ClientDocumentService.get_recent_documents(request.user, limit=5)
    
    # Get all documents for the sidebar links
    all_quotes = ClientDocumentService.get_accessible_quotes(request.user).order_by("-issue_date")
    all_invoices = ClientDocumentService.get_accessible_invoices(request.user).order_by("-issue_date")
    all_portal_documents = ClientDocumentService.get_accessible_portal_documents(request.user)
    client_tasks = ClientService.get_portal_tasks(client_record)
    
    return TemplateResponse(
        request,
        "core/client_dashboard_premium.html",
        {
            "client_record": client_record,
            "pending_quotes": recent_docs.get('quotes', []),
            "unpaid_invoices": recent_docs.get('invoices', []),
            "recent_portal_documents": recent_docs.get('portal_documents', []),
            "all_quotes": all_quotes,
            "all_invoices": all_invoices,
            "stats": stats,
            "quote_summary": _build_quote_summary(all_quotes),
            "invoice_summary": _build_invoice_summary(all_invoices),
            "document_summary": _build_document_summary(all_portal_documents),
            "recent_client_tasks": client_tasks[:4],
            "task_summary": ClientService.get_portal_task_summary(client_tasks),
        },
    )


@client_portal_required
def client_tasks(request):
    """Client portal view exposing the tasks linked to the authenticated client."""
    from .services.client_service import ClientService

    client_record = _get_client_portal_record(request.user)
    tasks = ClientService.get_portal_tasks(client_record)
    full_summary = ClientService.get_portal_task_summary(tasks)
    status_filter = request.GET.get('status', '').strip()

    if status_filter and status_filter in dict(Task.STATUS_CHOICES):
        tasks = tasks.filter(status=status_filter)

    return TemplateResponse(
        request,
        'core/client_tasks.html',
        {
            'client_record': client_record,
            'tasks': tasks,
            'task_summary': full_summary,
            'status_filter': status_filter,
        },
    )


@client_portal_required
def client_quotes(request):
    """Client Portal quotes list view."""
    quotes = ClientDocumentService.get_accessible_quotes(request.user).order_by("-issue_date")
    
    return TemplateResponse(
        request,
        "core/client_quotes_premium.html",
        {
            "client_record": _get_client_portal_record(request.user),
            "quotes": quotes,
            "quote_summary": _build_quote_summary(quotes),
        },
    )


@client_portal_required
def client_invoices(request):
    """Client Portal invoices list view."""
    invoices = ClientDocumentService.get_accessible_invoices(request.user).order_by("-issue_date")
    
    return TemplateResponse(
        request,
        "core/client_invoices_premium.html",
        {
            "client_record": _get_client_portal_record(request.user),
            "invoices": invoices,
            "invoice_summary": _build_invoice_summary(invoices),
        },
    )


@client_portal_required
def client_documents(request):
    """Client Portal manually published documents list."""
    documents = ClientDocumentService.get_accessible_portal_documents(request.user)

    return TemplateResponse(
        request,
        "core/client_documents_premium.html",
        {
            "client_record": _get_client_portal_record(request.user),
            "documents": documents,
            "document_summary": _build_document_summary(documents),
        },
    )


@client_portal_required
def client_submit_document(request):
    """Client Portal: soumettre un document à l'équipe NetExpress."""
    from .forms import ClientDocumentUploadForm
    from .models import ClientSubmittedDocument

    if request.method == 'POST':
        form = ClientDocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.client_user = request.user
            doc.save()
            messages.success(request, f"Document « {doc.title} » envoyé avec succès. Notre équipe le traitera rapidement.")
            return redirect('core:client_documents')
    else:
        form = ClientDocumentUploadForm()

    # Historique des documents soumis par ce client
    submitted = ClientSubmittedDocument.objects.filter(client_user=request.user).order_by('-submitted_at')[:20]

    return TemplateResponse(
        request,
        "core/client_submit_document.html",
        {"form": form, "submitted_documents": submitted},
    )


@client_portal_required
def client_document_detail(request, pk):
    """Client Portal document detail view."""
    from .models import ClientPortalDocument

    document = get_object_or_404(
        ClientPortalDocument.objects.select_related('client', 'published_by'),
        pk=pk,
    )

    if not ClientDocumentService.can_access_portal_document(request.user, document):
        return TemplateResponse(request, "core/access_denied.html", {"target": "document"}, status=403)

    ClientDocumentService.track_portal_document_access(document)

    return TemplateResponse(
        request,
        "core/client_document_detail_premium.html",
        {"document": document},
    )


@client_portal_required
def client_document_download(request, pk):
    """Secure download endpoint for published client portal documents."""
    import mimetypes
    from django.http import FileResponse, Http404
    from .models import ClientPortalDocument

    document = get_object_or_404(ClientPortalDocument, pk=pk)
    if not ClientDocumentService.can_access_portal_document(request.user, document):
        return TemplateResponse(request, "core/access_denied.html", {"target": "document"}, status=403)

    if not document.file:
        raise Http404("Le fichier demandé est introuvable.")

    inline = request.GET.get('inline') == '1'
    ClientDocumentService.track_portal_document_access(document, download=not inline)
    content_type, _ = mimetypes.guess_type(document.file.name)
    response = FileResponse(document.file.open('rb'), content_type=content_type or 'application/octet-stream')
    response['Content-Disposition'] = (
        f"inline; filename=\"{document.filename}\"" if inline else f"attachment; filename=\"{document.filename}\""
    )
    return response


@client_portal_required
def client_quote_detail(request, pk):
    """Client Portal quote detail view."""
    # Get the quote and check access
    quote = get_object_or_404(Quote, pk=pk)
    
    if not ClientDocumentService.can_access_quote(request.user, quote):
        return TemplateResponse(request, "core/access_denied.html", {"target": "quote"}, status=403)
    
    # Track document access
    ClientDocumentService.track_document_access(request.user, quote=quote)
    
    return TemplateResponse(
        request,
        "core/client_quote_detail_premium.html",
        {"quote": quote},
    )


@client_portal_required
def client_invoice_detail(request, pk):
    """Client Portal invoice detail view."""
    # Get the invoice and check access
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if not ClientDocumentService.can_access_invoice(request.user, invoice):
        return TemplateResponse(request, "core/access_denied.html", {"target": "invoice"}, status=403)
    
    # Track document access
    ClientDocumentService.track_document_access(request.user, invoice=invoice)
    
    return TemplateResponse(
        request,
        "core/client_invoice_detail_premium.html",
        {"invoice": invoice},
    )


@client_portal_required
def client_invoice_download(request, pk):
    """Client Portal invoice PDF download."""
    from django.http import FileResponse, Http404
    
    # Get the invoice and check access
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if not ClientDocumentService.can_access_invoice(request.user, invoice):
        return TemplateResponse(request, "core/access_denied.html", {"target": "invoice"}, status=403)
    
    # Generate PDF if not exists
    if not invoice.pdf:
        try:
            invoice.generate_pdf(attach=True)
        except Exception:
            raise Http404("Cette facture n'a pas de PDF et sa génération a échoué.")
    
    # Track document access
    ClientDocumentService.track_document_access(request.user, invoice=invoice)
    
    response = FileResponse(
        invoice.pdf.open("rb"),
        content_type="application/pdf"
    )
    response['Content-Disposition'] = f'inline; filename="facture_{invoice.number}.pdf"'
    return response


@client_portal_required
def client_quote_download(request, pk):
    """Client Portal quote PDF download."""
    from django.http import FileResponse, Http404
    
    # Get the quote and check access
    quote = get_object_or_404(Quote, pk=pk)
    
    if not ClientDocumentService.can_access_quote(request.user, quote):
        return TemplateResponse(request, "core/access_denied.html", {"target": "quote"}, status=403)
    
    # Generate PDF if not exists
    if not quote.pdf:
        try:
            quote.generate_pdf(attach=True)
        except Exception:
            raise Http404("Ce devis n'a pas de PDF et sa génération a échoué.")
    
    # Track document access
    ClientDocumentService.track_document_access(request.user, quote=quote)
    
    response = FileResponse(
        quote.pdf.open("rb"),
        content_type="application/pdf"
    )
    response['Content-Disposition'] = f'inline; filename="devis_{quote.number}.pdf"'
    return response


@client_portal_required
def client_quote_validate_start(request, pk):
    """Client Portal quote validation initiation."""
    # Get the quote and check access
    quote = get_object_or_404(Quote, pk=pk)
    
    if not ClientDocumentService.can_access_quote(request.user, quote):
        return TemplateResponse(request, "core/access_denied.html", {"target": "quote"}, status=403)
    
    # Create validation and send code (use-case layer)
    from devis.application.quote_validation import QuoteNotValidatableError, start_quote_validation

    try:
        res = start_quote_validation(quote, request=request)
        validation = res.validation
    except QuoteNotValidatableError:
        return TemplateResponse(
            request,
            "core/client_quote_detail_premium.html",
            {"quote": quote, "error_message": "Ce devis ne peut pas être validé dans son état actuel."},
        )
    
    return TemplateResponse(
        request,
        "core/client_quote_validate_code.html",
        {
            "quote": quote,
            "validation": validation,
            "success_message": "Un code de validation a été envoyé à votre adresse email."
        },
    )


@client_portal_required
def client_quote_validate_code(request, pk):
    """Client Portal quote validation code verification."""
    # Get the quote and check access
    quote = get_object_or_404(Quote, pk=pk)
    
    if not ClientDocumentService.can_access_quote(request.user, quote):
        return TemplateResponse(request, "core/access_denied.html", {"target": "quote"}, status=403)
    
    # Get the most recent validation for this quote (use-case helper)
    from devis.application.quote_validation import get_pending_validation_for_quote
    from devis.models import QuoteValidation

    try:
        validation = get_pending_validation_for_quote(quote)
    except QuoteValidation.DoesNotExist:
        return TemplateResponse(
            request,
            "core/client_quote_detail_premium.html",
            {
                "quote": quote,
                "error_message": "Aucune validation en cours pour ce devis."
            }
        )
    
    if validation.is_expired:
        return TemplateResponse(
            request,
            "core/client_quote_validate_expired.html",
            {"quote": quote, "validation": validation}
        )
    
    if request.method == "POST":
        from devis.forms import QuoteValidationCodeForm
        from devis.application.quote_validation import QuoteValidationExpiredError, confirm_quote_validation_code

        form = QuoteValidationCodeForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data["code"]
            try:
                ok = confirm_quote_validation_code(validation=validation, submitted_code=code)
            except QuoteValidationExpiredError:
                return TemplateResponse(
                    request,
                    "core/client_quote_validate_expired.html",
                    {"quote": quote, "validation": validation},
                )

            if ok:
                return TemplateResponse(
                    request,
                    "core/client_quote_validate_success.html",
                    {"quote": quote},
                )

            form.add_error("code", "Code incorrect. Veuillez réessayer.")
    else:
        from devis.forms import QuoteValidationCodeForm
        form = QuoteValidationCodeForm()
    
    return TemplateResponse(
        request,
        "core/client_quote_validate_code.html",
        {
            "quote": quote,
            "validation": validation,
            "form": form,
        },
    )


@worker_portal_required
def worker_dashboard(request):
    """Tableau de bord ouvrier (tâches assignées à l'utilisateur connecté)."""
    # Filter tasks by assigned worker (new approach)
    user_tasks = Task.objects.filter(assigned_to=request.user).order_by('due_date', 'start_date')
    
    # Separate tasks by status for better organization
    upcoming_tasks = user_tasks.filter(status=Task.STATUS_UPCOMING)
    in_progress_tasks = user_tasks.filter(status=Task.STATUS_IN_PROGRESS)
    almost_overdue_tasks = user_tasks.filter(status=Task.STATUS_ALMOST_OVERDUE)
    overdue_tasks = user_tasks.filter(status=Task.STATUS_OVERDUE)
    completed_tasks = user_tasks.filter(status=Task.STATUS_COMPLETED)[:10]  # Show last 10 completed
    
    return render(
        request,
        "core/worker_dashboard.html",
        {
            "upcoming_tasks": upcoming_tasks,
            "in_progress_tasks": in_progress_tasks,
            "almost_overdue_tasks": almost_overdue_tasks,
            "overdue_tasks": overdue_tasks,
            "completed_tasks": completed_tasks,
            "total_assigned": user_tasks.count(),
            "pending_tasks": user_tasks.exclude(status=Task.STATUS_COMPLETED).count(),
        },
    )


@admin_portal_required
def admin_dashboard(request):
    """Admin Portal dashboard with KPIs and performance metrics."""
    from decimal import Decimal
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.contrib.auth.models import User
    from django.core.cache import cache
    import json
    
    # Cache key for dashboard data (5 minutes cache)
    cache_key = 'admin_dashboard_data'
    
    # Permettre de forcer le rafraîchissement via le bouton "Actualiser"
    if request.GET.get('refresh') == '1':
        cache.delete(cache_key)
    
    cached_data = cache.get(cache_key)
    
    if cached_data is None:
        # Calculate revenue metrics with optimized queries
        total_revenue = Invoice.objects.filter(
            status__in=[Invoice.InvoiceStatus.PAID, Invoice.InvoiceStatus.PARTIAL]
        ).aggregate(total=Sum('total_ttc'))['total'] or Decimal('0.00')
        
        # Monthly revenue (current month)
        current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        monthly_revenue = Invoice.objects.filter(
            status__in=[Invoice.InvoiceStatus.PAID, Invoice.InvoiceStatus.PARTIAL],
            issue_date__gte=current_month
        ).aggregate(total=Sum('total_ttc'))['total'] or Decimal('0.00')
        
        # Pending revenue (sent but not paid invoices)
        pending_revenue = Invoice.objects.filter(
            status=Invoice.InvoiceStatus.SENT
        ).aggregate(total=Sum('total_ttc'))['total'] or Decimal('0.00')
        
        # Quote conversion rate - optimized with single query
        quote_stats = Quote.objects.aggregate(
            total=Count('id'),
            accepted=Count('id', filter=Q(status=Quote.QuoteStatus.ACCEPTED))
        )
        total_quotes = quote_stats['total']
        accepted_quotes = quote_stats['accepted']
        conversion_rate = (accepted_quotes / total_quotes * 100) if total_quotes > 0 else 0
        
        # Task performance metrics - optimized
        task_stats = Task.objects.aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status=Task.STATUS_COMPLETED)),
            overdue=Count('id', filter=Q(status=Task.STATUS_OVERDUE))
        )
        total_tasks = task_stats['total']
        completed_tasks = task_stats['completed']
        overdue_tasks = task_stats['overdue']
        task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        # Worker performance
        # NOTE: Task.assigned_to utilise related_name='assigned_tasks' (pas de 'task_set').
        # On évite donc un prefetch invalide qui casserait le dashboard.
        workers = User.objects.filter(groups__name='Workers')
        worker_stats = []
        for worker in workers:
            worker_tasks = Task.objects.filter(assigned_to=worker)
            completed = worker_tasks.filter(status=Task.STATUS_COMPLETED).count()
            total = worker_tasks.count()
            completion_rate = (completed / total * 100) if total > 0 else 0
            worker_stats.append({
                'worker': worker,
                'total_tasks': total,
                'completed_tasks': completed,
                'completion_rate': completion_rate,
                'overdue_tasks': worker_tasks.filter(status=Task.STATUS_OVERDUE).count()
            })
        
        # Recent activity - optimized with select_related
        recent_quotes = Quote.objects.select_related('client', 'service').order_by('-created_at')[:5]
        recent_invoices = Invoice.objects.select_related('quote', 'quote__client').order_by('-created_at')[:5]
        recent_tasks = Task.objects.prefetch_related('assigned_to').order_by('-created_at')[:5]
        
        # Status distributions - optimized with single aggregation
        quote_status_counts = dict(
            Quote.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        quote_status_counts = {
            'draft': quote_status_counts.get(Quote.QuoteStatus.DRAFT, 0),
            'sent': quote_status_counts.get(Quote.QuoteStatus.SENT, 0),
            'accepted': quote_status_counts.get(Quote.QuoteStatus.ACCEPTED, 0),
            'rejected': quote_status_counts.get(Quote.QuoteStatus.REJECTED, 0),
        }
        
        invoice_status_counts = dict(
            Invoice.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        invoice_status_counts = {
            'draft': invoice_status_counts.get(Invoice.InvoiceStatus.DRAFT, 0),
            'sent': invoice_status_counts.get(Invoice.InvoiceStatus.SENT, 0),
            'paid': invoice_status_counts.get(Invoice.InvoiceStatus.PAID, 0),
            'overdue': invoice_status_counts.get(Invoice.InvoiceStatus.OVERDUE, 0),
        }
        
        task_status_counts = dict(
            Task.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        task_status_counts = {
            'upcoming': task_status_counts.get(Task.STATUS_UPCOMING, 0),
            'in_progress': task_status_counts.get(Task.STATUS_IN_PROGRESS, 0),
            'completed': task_status_counts.get(Task.STATUS_COMPLETED, 0),
            'overdue': task_status_counts.get(Task.STATUS_OVERDUE, 0),
        }
        
        # Revenue trend (last 6 months) - optimized
        revenue_trend = []
        for i in range(6):
            month_start = (current_month - timedelta(days=30*i)).replace(day=1)
            if i == 0:
                month_end = timezone.now().date()
            else:
                month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            month_revenue = Invoice.objects.filter(
                status__in=[Invoice.InvoiceStatus.PAID, Invoice.InvoiceStatus.PARTIAL],
                issue_date__gte=month_start,
                issue_date__lte=month_end
            ).aggregate(total=Sum('total_ttc'))['total'] or Decimal('0.00')
            revenue_trend.append({
                'month': month_start.strftime('%Y-%m'),
                'revenue': float(month_revenue)
            })
        revenue_trend.reverse()
        
        # Store only serializable data in cache
        cached_data = {
            'total_revenue': float(total_revenue),
            'monthly_revenue': float(monthly_revenue),
            'pending_revenue': float(pending_revenue),
            'conversion_rate': float(conversion_rate),
            'task_completion_rate': float(task_completion_rate),
            'worker_stats': [
                {
                    'worker_id': ws['worker'].id,
                    'worker_name': ws['worker'].get_full_name() or ws['worker'].username,
                    'total_tasks': ws['total_tasks'],
                    'completed_tasks': ws['completed_tasks'],
                    'completion_rate': float(ws['completion_rate']),
                    'overdue_tasks': ws['overdue_tasks']
                }
                for ws in worker_stats
            ],
            'recent_quotes_ids': [q.id for q in recent_quotes],
            'recent_invoices_ids': [i.id for i in recent_invoices],
            'recent_tasks_ids': [t.id for t in recent_tasks],
            'quote_status_counts': quote_status_counts,
            'invoice_status_counts': invoice_status_counts,
            'task_status_counts': task_status_counts,
            'total_quotes': total_quotes,
            'total_invoices': Invoice.objects.count(),
            'total_tasks': total_tasks,
            'overdue_tasks': overdue_tasks,
            'revenue_trend': revenue_trend,
            'revenue_trend_json': json.dumps(revenue_trend),
        }
        cache.set(cache_key, cached_data, 300)  # Cache for 5 minutes
        
        # Keep querysets for template rendering
        recent_quotes_qs = recent_quotes
        recent_invoices_qs = recent_invoices
        recent_tasks_qs = recent_tasks
        worker_stats_list = worker_stats
        revenue_trend_json = cached_data['revenue_trend_json']
    else:
        # Restore querysets from cached IDs
        recent_quotes_qs = Quote.objects.filter(
            pk__in=cached_data['recent_quotes_ids']
        ).select_related('client', 'service').order_by('-created_at')
        recent_invoices_qs = Invoice.objects.filter(
            pk__in=cached_data['recent_invoices_ids']
        ).select_related('quote', 'quote__client').order_by('-created_at')
        recent_tasks_qs = Task.objects.filter(
            pk__in=cached_data['recent_tasks_ids']
        ).prefetch_related('assigned_to').order_by('-created_at')
        
        # Restore worker stats with user objects
        worker_stats_list = []
        for ws_data in cached_data['worker_stats']:
            try:
                worker = User.objects.get(pk=ws_data['worker_id'])
                worker_stats_list.append({
                    'worker': worker,
                    'total_tasks': ws_data['total_tasks'],
                    'completed_tasks': ws_data['completed_tasks'],
                    'completion_rate': ws_data['completion_rate'],
                    'overdue_tasks': ws_data['overdue_tasks']
                })
            except User.DoesNotExist:
                continue
        
        # Get revenue trend JSON
        revenue_trend_json = cached_data.get('revenue_trend_json', json.dumps(cached_data.get('revenue_trend', [])))
    
    return render(
        request,
        "core/admin_dashboard.html",
        {
            # KPI Metrics
            'total_revenue': Decimal(str(cached_data['total_revenue'])),
            'monthly_revenue': Decimal(str(cached_data['monthly_revenue'])),
            'pending_revenue': Decimal(str(cached_data['pending_revenue'])),
            'conversion_rate': cached_data['conversion_rate'],
            'task_completion_rate': cached_data['task_completion_rate'],
            
            # Worker Performance
            'worker_stats': worker_stats_list,
            'total_workers': len(worker_stats_list),
            
            # Recent Activity
            'recent_quotes': recent_quotes_qs,
            'recent_invoices': recent_invoices_qs,
            'recent_tasks': recent_tasks_qs,
            
            # Status Distributions
            'quote_status_counts': cached_data['quote_status_counts'],
            'invoice_status_counts': cached_data['invoice_status_counts'],
            'task_status_counts': cached_data['task_status_counts'],
            
            # Totals
            'total_quotes': cached_data['total_quotes'],
            'total_invoices': cached_data['total_invoices'],
            'total_tasks': cached_data['total_tasks'],
            'overdue_tasks': cached_data['overdue_tasks'],
            
            # Revenue Trend for Charts (as JSON string for template)
            'revenue_trend': cached_data['revenue_trend'],
            'revenue_trend_json': revenue_trend_json,
        },
    )


@admin_portal_required
def admin_email_health(request):
    """Centre de supervision email pour le portail admin."""
    from .services.email_health_service import EmailHealthService

    email_health = EmailHealthService.get_configuration_report()
    probe_result = None
    test_result = None
    default_test_recipient = request.user.email or ""

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'probe':
            probe_result = EmailHealthService.probe_delivery_backend()
            if probe_result['status'] == 'healthy':
                messages.success(request, probe_result['title'])
            elif probe_result['status'] == 'warning':
                messages.warning(request, probe_result['title'])
            else:
                messages.error(request, probe_result['title'])
        elif action == 'send_test_email':
            default_test_recipient = (request.POST.get('recipient_email') or '').strip()
            test_result = EmailHealthService.send_test_email(default_test_recipient, triggered_by=request.user)
            if test_result['status'] == 'healthy':
                messages.success(request, test_result['title'])
            elif test_result['status'] == 'warning':
                messages.warning(request, test_result['title'])
            else:
                messages.error(request, test_result['title'])
        else:
            messages.warning(request, "Action de supervision email inconnue.")

        email_health = EmailHealthService.get_configuration_report()

    return render(
        request,
        'core/admin_email_health.html',
        {
            'email_health': email_health,
            'probe_result': probe_result,
            'test_result': test_result,
            'default_test_recipient': default_test_recipient,
        },
    )


@admin_portal_required
def admin_global_planning(request):
    """Admin Portal global planning view showing all workers and tasks."""
    from django.contrib.auth.models import User
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Get all workers
    workers = User.objects.filter(groups__name='Workers').order_by('first_name', 'last_name')
    
    # Get date range for planning (default to current week)
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    # Allow filtering by date range via GET parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = start_of_week
    else:
        start_date = start_of_week
        
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = end_of_week
    else:
        end_date = end_of_week
    
    # Get all tasks in the date range
    all_tasks = Task.objects.filter(
        start_date__lte=end_date,
        due_date__gte=start_date
    ).order_by('start_date', 'due_date')
    
    # Organize tasks by worker
    worker_planning = []
    for worker in workers:
        worker_tasks = all_tasks.filter(assigned_to=worker)
        
        # Categorize tasks by status
        upcoming = worker_tasks.filter(status=Task.STATUS_UPCOMING)
        in_progress = worker_tasks.filter(status=Task.STATUS_IN_PROGRESS)
        completed = worker_tasks.filter(status=Task.STATUS_COMPLETED)
        overdue = worker_tasks.filter(status=Task.STATUS_OVERDUE)
        almost_overdue = worker_tasks.filter(status=Task.STATUS_ALMOST_OVERDUE)
        
        worker_planning.append({
            'worker': worker,
            'total_tasks': worker_tasks.count(),
            'upcoming_tasks': upcoming,
            'in_progress_tasks': in_progress,
            'completed_tasks': completed,
            'overdue_tasks': overdue,
            'almost_overdue_tasks': almost_overdue,
            'workload_percentage': min(100, (worker_tasks.count() * 10))  # Simple workload calculation
        })
    
    # Get unassigned tasks (ManyToMany requires annotation)
    from django.db.models import Count as CountAnnotate
    all_tasks_annotated = all_tasks.annotate(worker_count=CountAnnotate('assigned_to'))
    unassigned_tasks = all_tasks_annotated.filter(worker_count=0)
    
    # Summary statistics
    total_tasks_in_period = all_tasks.count()
    assigned_tasks_count = all_tasks_annotated.filter(worker_count__gt=0).count()
    unassigned_count = unassigned_tasks.count()
    
    # Task distribution by status
    status_distribution = {
        'upcoming': all_tasks.filter(status=Task.STATUS_UPCOMING).count(),
        'in_progress': all_tasks.filter(status=Task.STATUS_IN_PROGRESS).count(),
        'completed': all_tasks.filter(status=Task.STATUS_COMPLETED).count(),
        'overdue': all_tasks.filter(status=Task.STATUS_OVERDUE).count(),
        'almost_overdue': all_tasks.filter(status=Task.STATUS_ALMOST_OVERDUE).count(),
    }
    
    return render(
        request,
        "core/admin_global_planning.html",
        {
            'worker_planning': worker_planning,
            'unassigned_tasks': unassigned_tasks,
            'start_date': start_date,
            'end_date': end_date,
            'total_workers': workers.count(),
            'total_tasks_in_period': total_tasks_in_period,
            'assigned_tasks': assigned_tasks_count,
            'unassigned_count': unassigned_count,
            'status_distribution': status_distribution,
            'all_tasks': all_tasks,  # For comprehensive view
        },
    )


# Admin Portal Management Views

@admin_portal_required
def admin_create_worker(request):
    """Admin Portal worker creation view."""
    from .forms import WorkerCreationForm
    from .services.worker_service import WorkerService
    from django.core.exceptions import ValidationError
    
    if request.method == 'POST':
        form = WorkerCreationForm(request.POST)
        if form.is_valid():
            try:
                # Utiliser le service métier pour créer le worker
                # Note: On ignore password1/password2 du formulaire, le service génère un mot de passe temporaire
                worker, temporary_password = WorkerService.create_worker(
                    email=form.cleaned_data['email'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    phone=form.cleaned_data.get('phone', ''),
                    admin_user=request.user
                )
                
                # Envoyer les identifiants par email
                email_sent = WorkerService.send_worker_credentials(worker, temporary_password, request)
                
                if email_sent:
                    messages.success(
                        request,
                        f"Ouvrier {worker.get_full_name()} créé avec succès! "
                        f"Les identifiants ont été envoyés par email."
                    )
                else:
                    messages.warning(
                        request,
                        f"Ouvrier {worker.get_full_name()} créé avec succès, "
                        f"mais l'email d'identifiants n'a pas pu être envoyé. "
                        f"Mot de passe temporaire : {temporary_password}"
                    )
                
                return redirect('core:admin_worker_detail', pk=worker.pk)
            except ValidationError as e:
                form.add_error(None, str(e))
    else:
        form = WorkerCreationForm()
    
    return render(request, 'core/admin_create_worker.html', {'form': form})


@admin_portal_required
def admin_create_client(request):
    """Admin Portal client creation view."""
    from .forms import ClientCreationForm
    from .services.client_service import ClientService
    from accounts.services import ClientAccountCreationService
    from core.services.email_service import EmailService
    from django.core.exceptions import ValidationError
    
    if request.method == 'POST':
        form = ClientCreationForm(request.POST)
        if form.is_valid():
            try:
                # Utiliser le service métier pour créer le client
                client = ClientService.create_client(
                    full_name=form.cleaned_data['full_name'],
                    email=form.cleaned_data['email'],
                    phone=form.cleaned_data['phone'],
                    address_line=form.cleaned_data.get('address_line', ''),
                    city=form.cleaned_data.get('city', ''),
                    zip_code=form.cleaned_data.get('zip_code', ''),
                    company=form.cleaned_data.get('company', ''),
                    link_to_user=False  # Option pour lier à un User existant (futur)
                )
                messages.success(request, f"Client {client.full_name} créé avec succès!")

                if form.cleaned_data.get('create_portal_access'):
                    portal_user, account_created = ClientAccountCreationService.create_from_client(client)
                    invitation_sent = EmailService.send_client_portal_invitation(portal_user, client, request=request)
                    if invitation_sent:
                        status_label = "activé" if account_created else "mis à jour"
                        messages.success(request, f"Espace client {status_label} et invitation envoyée à {portal_user.email}.")
                    else:
                        messages.warning(
                            request,
                            "Le compte portail a été préparé, mais l'invitation email n'a pas pu être envoyée. "
                            "Vérifiez la configuration mail ou utilisez le flux mot de passe oublié."
                        )

                return redirect('core:admin_client_detail', pk=client.pk)
            except ValidationError as e:
                form.add_error(None, str(e))
    else:
        form = ClientCreationForm()
    
    return render(request, 'core/admin_create_client.html', {'form': form})


@admin_portal_required
def admin_create_quote(request):
    """Admin Portal quote creation view with inline items (like Django Admin)."""
    from .forms import QuoteCreationForm, QuoteItemFormset
    from services.models import Service
    
    # Get services for JavaScript autofill (Service model has no price field)
    services_data = list(Service.objects.filter(is_active=True).values('id', 'title'))
    prefilled_client = None

    if request.method == 'GET' and request.GET.get('client'):
        prefilled_client = Client.objects.filter(pk=request.GET.get('client')).first()
    
    if request.method == 'POST':
        form = QuoteCreationForm(request.POST)
        formset = QuoteItemFormset(request.POST, prefix='items')
        if request.POST.get('client'):
            prefilled_client = Client.objects.filter(pk=request.POST.get('client')).first()
        
        if form.is_valid() and formset.is_valid():
            quote = form.save()
            # Save formset items linked to the quote
            items = formset.save(commit=False)
            for item in items:
                item.quote = quote
                item.save()
            # Handle deleted items
            for obj in formset.deleted_objects:
                obj.delete()
            # Recalculate totals
            quote.compute_totals()
            messages.success(request, f"Devis {quote.number} créé avec succès!")
            return redirect('core:admin_quote_detail', pk=quote.pk)
    else:
        initial = {'client': prefilled_client.pk} if prefilled_client else None
        form = QuoteCreationForm(initial=initial)
        formset = QuoteItemFormset(prefix='items')
    
    import json
    return render(request, 'core/admin_create_quote.html', {
        'form': form,
        'formset': formset,
        'services_json': json.dumps(services_data, default=str),
        'prefilled_client': prefilled_client,
    })


@admin_portal_required
def admin_edit_quote(request, pk):
    """Admin Portal quote editing view with inline items (like Django Admin)."""
    from .forms import QuoteCreationForm, QuoteItemFormset
    from services.models import Service
    from devis.models import Quote, QuoteItem
    
    quote = get_object_or_404(Quote, pk=pk)
    
    # Get services for JavaScript autofill (Service model has no price field)
    services_data = list(Service.objects.filter(is_active=True).values('id', 'title'))
    
    if request.method == 'POST':
        form = QuoteCreationForm(request.POST, instance=quote)
        formset = QuoteItemFormset(request.POST, instance=quote, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            quote = form.save()
            
            # Handle deletions FIRST - before saving new/updated items
            # This ensures deleted items are removed even if formset.deleted_objects is empty
            for item_form in formset:
                if item_form.cleaned_data.get('DELETE') and item_form.instance.pk:
                    item_form.instance.delete()
            
            # Save formset (new and updated items only)
            instances = formset.save(commit=False)
            for instance in instances:
                instance.quote = quote
                instance.save()
            
            # Also handle deleted_objects as a fallback (for compatibility)
            for obj in formset.deleted_objects:
                if obj.pk:  # Only delete if it exists in DB
                    obj.delete()
            
            # Recalculate totals
            quote.compute_totals()
            messages.success(request, f"Devis {quote.number} mis à jour avec succès!")
            return redirect('core:admin_quote_detail', pk=quote.pk)
    else:
        form = QuoteCreationForm(instance=quote)
        formset = QuoteItemFormset(instance=quote, prefix='items')
    
    import json
    return render(request, 'core/admin_edit_quote.html', {
        'form': form,
        'formset': formset,
        'quote': quote,
        'services_json': json.dumps(services_data, default=str),
    })


@admin_portal_required
def admin_create_task(request):
    """Admin Portal task creation view."""
    from django.conf import settings
    from django.contrib.auth.models import User
    from .forms import TaskCreationForm

    worker_id = request.GET.get('worker')
    client_id = request.GET.get('client')
    prefilled_worker = None
    prefilled_client = None
    initial = {}

    if worker_id:
        prefilled_worker = User.objects.filter(pk=worker_id, groups__name='Workers').first()
        if prefilled_worker:
            initial['assigned_to'] = [prefilled_worker.pk]

    if client_id:
        prefilled_client = Client.objects.filter(pk=client_id).first()
        if prefilled_client:
            initial['client'] = prefilled_client.pk
            initial.setdefault('title', f"Suivi client - {prefilled_client.company or prefilled_client.full_name}")
            location_parts = [
                prefilled_client.address_line,
                " ".join(part for part in [prefilled_client.zip_code, prefilled_client.city] if part).strip(),
            ]
            prefilled_location = ", ".join(part for part in location_parts if part)
            if prefilled_location:
                initial.setdefault('location', prefilled_location)
            if prefilled_client.company:
                initial.setdefault('team', prefilled_client.company)

    branding = getattr(settings, 'INVOICE_BRANDING', {})
    
    if request.method == 'POST':
        form = TaskCreationForm(request.POST)
        if form.is_valid():
            task = form.save()
            messages.success(request, f"Tâche '{task.title}' créée avec succès!")
            return redirect('core:admin_task_detail', pk=task.pk)
    else:
        form = TaskCreationForm(initial=initial)
    
    return render(request, 'core/admin_create_task.html', {
        'form': form,
        'company_name': branding.get('name', 'NetExpress'),
        'company_siret': branding.get('siret', ''),
        'prefilled_worker': prefilled_worker,
        'prefilled_client': prefilled_client,
    })


@admin_portal_required
def admin_edit_task(request, pk):
    """Admin Portal task edit view."""
    from .forms import TaskCreationForm
    from tasks.models import Task
    
    task = get_object_or_404(Task, pk=pk)
    
    if request.method == 'POST':
        form = TaskCreationForm(request.POST, instance=task)
        if form.is_valid():
            task = form.save()
            messages.success(request, f"Tâche '{task.title}' mise à jour avec succès!")
            return redirect('core:admin_task_detail', pk=task.pk)
    else:
        form = TaskCreationForm(instance=task)
    
    return render(request, 'core/admin_edit_task.html', {'form': form, 'task': task})


@admin_portal_required
def admin_delete_task(request, pk):
    """Admin Portal task deletion view."""
    from tasks.models import Task
    
    task = get_object_or_404(Task, pk=pk)
    
    if request.method == 'POST':
        title = task.title
        task.delete()
        messages.success(request, f"Tâche '{title}' supprimée avec succès!")
        return redirect('core:admin_tasks_list')
    
    return render(request, 'core/admin_delete_task.html', {'task': task})


@admin_portal_required
def admin_task_change_status(request, pk):
    """Admin Portal task status change view (HTMX compatible)."""
    from tasks.models import Task
    
    task = get_object_or_404(Task, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Task.STATUS_CHOICES):
            old_status = task.status
            task.status = new_status
            # Si marqué comme terminé, enregistrer qui l'a fait
            if new_status == Task.STATUS_COMPLETED and old_status != Task.STATUS_COMPLETED:
                task.completed_by = request.user
            task.save()
            messages.success(request, f"Statut de '{task.title}' changé en '{task.get_status_display()}'")
        else:
            messages.error(request, "Statut invalide")
    
    # Rediriger vers la page d'origine
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('core:admin_task_detail', pk=pk)


@admin_portal_required
def admin_task_mark_complete(request, pk):
    """Admin Portal quick action to mark task as complete."""
    from tasks.models import Task
    
    task = get_object_or_404(Task, pk=pk)
    
    if request.method == 'POST':
        if task.status != Task.STATUS_COMPLETED:
            task.status = Task.STATUS_COMPLETED
            task.completed_by = request.user
            task.completion_notes = request.POST.get('completion_notes', '')
            task.save()
            messages.success(request, f"Tâche '{task.title}' marquée comme terminée!")
        else:
            messages.info(request, f"La tâche '{task.title}' est déjà terminée.")
    
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('core:admin_task_detail', pk=pk)


@admin_portal_required
def admin_workers_list(request):
    """Admin Portal workers management view with optimized queries."""
    from django.contrib.auth.models import User
    from django.db.models import Count, Q
    from django.core.paginator import Paginator
    from accounts.models import Profile
    
    # Filtrer par rôle worker dans le profil
    workers = User.objects.filter(
        profile__role=Profile.ROLE_WORKER
    ).select_related('profile').order_by('first_name', 'last_name')
    
    # Optimize with annotations (noms uniques pour éviter conflits)
    workers = workers.annotate(
        tasks_total=Count('assigned_tasks'),
        tasks_done=Count('assigned_tasks', filter=Q(assigned_tasks__status=Task.STATUS_COMPLETED)),
        tasks_late=Count('assigned_tasks', filter=Q(assigned_tasks__status=Task.STATUS_OVERDUE))
    )
    
    # Add task statistics for each worker
    workers_with_stats = []
    for worker in workers:
        workers_with_stats.append({
            'worker': worker,
            'total_tasks': worker.tasks_total,
            'completed_tasks': worker.tasks_done,
            'pending_tasks': worker.tasks_total - worker.tasks_done,
            'overdue_tasks': worker.tasks_late,
        })
    
    # Pagination
    paginator = Paginator(workers_with_stats, 20)  # 20 workers per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/admin_workers_list.html', {
        'workers_with_stats': page_obj,
        'page_obj': page_obj
    })


@admin_portal_required
def admin_clients_list(request):
    """Admin Portal clients management view with optimized queries."""
    from django.db.models import Count, Q, Sum, Exists, OuterRef, Max, Subquery, DecimalField, Value
    from django.db.models.functions import Coalesce
    from django.core.paginator import Paginator
    from factures.models import Invoice
    from .models import ClientPortalDocument
    
    # Get search and filter parameters
    search_query = request.GET.get('q', '').strip()
    filter_type = request.GET.get('filter', '')
    
    portal_user_subquery = User.objects.filter(email__iexact=OuterRef('email'))
    revenue_subquery = Invoice.objects.filter(
        quote__client=OuterRef('pk')
    ).values('quote__client').annotate(
        total=Sum('total_ttc')
    ).values('total')[:1]
    open_balance_subquery = Invoice.objects.filter(
        quote__client=OuterRef('pk'),
        status__in=[Invoice.InvoiceStatus.SENT, Invoice.InvoiceStatus.PARTIAL, Invoice.InvoiceStatus.OVERDUE],
    ).values('quote__client').annotate(
        total=Sum('total_ttc')
    ).values('total')[:1]
    clients = Client.objects.all()
    
    # Optimize with annotations - use 'quotes' (the related_name from Quote.client)
    clients = clients.annotate(
        total_quotes=Count('quotes', distinct=True),
        accepted_quotes=Count('quotes', filter=Q(quotes__status=Quote.QuoteStatus.ACCEPTED), distinct=True),
        pending_quotes=Count('quotes', filter=Q(quotes__status=Quote.QuoteStatus.SENT), distinct=True),
        total_invoices=Count('quotes__invoices', distinct=True),
        paid_invoices=Count('quotes__invoices', filter=Q(quotes__invoices__status=Invoice.InvoiceStatus.PAID), distinct=True),
        portal_enabled=Exists(portal_user_subquery),
        published_documents=Count('portal_documents', filter=Q(portal_documents__is_published=True), distinct=True),
        total_revenue=Coalesce(
            Subquery(revenue_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
            Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
        open_balance=Coalesce(
            Subquery(open_balance_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
            Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
        last_quote_date=Max('quotes__issue_date'),
    )
    
    # Search filter - search in name, email, phone, company
    if search_query:
        clients = clients.filter(
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(company__icontains=search_query) |
            Q(city__icontains=search_query)
        )
    
    # Status filter
    if filter_type == 'active':
        # Clients with at least one quote in the last 6 months
        six_months_ago = timezone.now() - timedelta(days=180)
        clients = clients.filter(quotes__created_at__gte=six_months_ago).distinct()
    elif filter_type == 'with_quotes':
        # Clients with at least one quote
        clients = clients.filter(total_quotes__gt=0)
    elif filter_type == 'no_quotes':
        # Clients without any quotes
        clients = clients.filter(total_quotes=0)
    elif filter_type == 'pending':
        # Clients with pending quotes
        clients = clients.filter(pending_quotes__gt=0)
    
    # Order by name
    clients = clients.order_by('full_name')
    filtered_total = clients.count()
    
    # Pagination
    paginator = Paginator(clients, 25)  # 25 clients per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    base_clients = Client.objects.annotate(portal_enabled=Exists(portal_user_subquery))
    six_months_ago = timezone.now() - timedelta(days=180)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    total_clients = Client.objects.count()
    clients_with_portal = base_clients.filter(portal_enabled=True).count()
    clients_without_quotes = Client.objects.annotate(total_quotes=Count('quotes')).filter(total_quotes=0).count()
    new_clients_30d = Client.objects.filter(created_at__gte=thirty_days_ago).count()
    portal_conversion_rate = round((clients_with_portal / total_clients) * 100, 1) if total_clients else 0
    
    return render(request, 'core/admin_clients_list.html', {
        'clients': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'filter_type': filter_type,
        'total_clients': total_clients,
        'filtered_total': filtered_total,
        'clients_with_portal': clients_with_portal,
        'clients_without_portal': max(total_clients - clients_with_portal, 0),
        'clients_without_quotes': clients_without_quotes,
        'new_clients_30d': new_clients_30d,
        'portal_conversion_rate': portal_conversion_rate,
        'active_clients': Client.objects.filter(quotes__created_at__gte=six_months_ago).distinct().count(),
        'clients_needing_followup': Client.objects.filter(
            quotes__invoices__status__in=[Invoice.InvoiceStatus.SENT, Invoice.InvoiceStatus.PARTIAL, Invoice.InvoiceStatus.OVERDUE]
        ).distinct().count(),
        'published_documents_total': ClientPortalDocument.objects.filter(is_published=True).count(),
    })


@admin_portal_required
def admin_client_import_template(request):
    """Download a ready-to-use Excel template for client imports."""
    from devis.import_service import IMPORT_COLUMN_GUIDE, TEMPLATE_HEADERS, TEMPLATE_SAMPLE_ROW

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        fallback_lines = [";".join(TEMPLATE_HEADERS), ";".join(TEMPLATE_SAMPLE_ROW)]
        response = HttpResponse("\n".join(fallback_lines), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="modele-import-clients-netexpress.csv"'
        return response

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clients"
    worksheet.append(TEMPLATE_HEADERS)
    worksheet.append(TEMPLATE_SAMPLE_ROW)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:G2"

    header_fill = PatternFill(fill_type="solid", fgColor="0E6B4C")
    note_fill = PatternFill(fill_type="solid", fgColor="F3F7F4")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    column_widths = [24, 32, 20, 32, 18, 16, 24]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width

    instructions = workbook.create_sheet("Consignes")
    instructions.append(["Colonne", "Obligatoire", "Noms acceptés"])
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for entry in IMPORT_COLUMN_GUIDE:
        instructions.append([
            entry["label"],
            "Oui" if entry["required"] else "Non",
            ", ".join(entry["aliases"]),
        ])

    instructions.append([])
    instructions.append(["Conseils", "", "L'email sert de clé de rapprochement pour mettre à jour un client existant."])
    instructions.append(["Conseils", "", "Les lignes totalement vides sont ignorées automatiquement."])
    instructions.append(["Conseils", "", "Les doublons d'email dans le même fichier sont signalés et ignorés."])

    for row in instructions.iter_rows(min_row=2, max_row=instructions.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row >= instructions.max_row - 2:
                cell.fill = note_fill

    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 64
    instructions.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="modele-import-clients-netexpress.xlsx"'
    return response


@admin_portal_required
def admin_import_clients(request):
    """Import clients from an Excel file (.xlsx)."""
    from devis.import_service import ALLOWED_EXTENSIONS, IMPORT_COLUMN_GUIDE, import_clients_from_excel

    result = None
    update_existing = request.method == 'POST' and request.POST.get('update_existing') == '1'
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Veuillez sélectionner un fichier Excel.")
        elif Path(excel_file.name).suffix.lower() not in ALLOWED_EXTENSIONS:
            messages.error(request, "Format invalide. Utilisez un fichier Excel .xlsx ou .xlsm.")
        elif excel_file.size > 5 * 1024 * 1024:  # 5 MB max
            messages.error(request, "Fichier trop volumineux (max 5 Mo).")
        else:
            result = import_clients_from_excel(excel_file, update_existing=update_existing)
            if result.created or result.updated:
                messages.success(
                    request,
                    f"Import terminé : {result.created} créé(s), {result.updated} mis à jour, {result.skipped} ignoré(s)."
                )
            elif result.errors:
                messages.error(request, "L'import a échoué. Voir les erreurs ci-dessous.")
            else:
                messages.info(request, "Aucun nouveau client à importer.")
    return render(request, 'core/admin_import_clients.html', {
        'accepted_extensions_text': ", ".join(sorted(ALLOWED_EXTENSIONS)),
        'import_column_guide': IMPORT_COLUMN_GUIDE,
        'max_file_size_mb': 5,
        'result': result,
        'update_existing': update_existing,
    })


@admin_portal_required
def admin_client_activate_portal_access(request, pk):
    """Create or refresh a client portal account from the admin dashboard."""
    from accounts.services import ClientAccountCreationService
    from core.services.email_service import EmailService
    from django.core.exceptions import ValidationError

    client = get_object_or_404(Client, pk=pk)
    if request.method != 'POST':
        return redirect('core:admin_client_detail', pk=client.pk)

    try:
        portal_user, account_created = ClientAccountCreationService.create_from_client(client)
        invitation_sent = EmailService.send_client_portal_invitation(portal_user, client, request=request)

        if invitation_sent:
            messages.success(
                request,
                f"Accès portail {'créé' if account_created else 'réactivé'} pour {client.full_name}. Invitation envoyée à {portal_user.email}."
            )
        else:
            messages.warning(
                request,
                "Le compte portail est prêt, mais l'invitation email n'a pas été envoyée. "
                "Contrôlez la configuration SMTP/Brevo puis renvoyez l'invitation."
            )
    except ValidationError as exc:
        messages.error(request, str(exc))

    return redirect('core:admin_client_detail', pk=client.pk)


@admin_portal_required
def admin_client_publish_document(request, pk):
    """Publish a document to a client's portal from the admin dashboard."""
    from django.urls import reverse
    from .forms import ClientPortalDocumentForm
    from .services.notification_service import notification_service

    client = get_object_or_404(Client, pk=pk)
    if request.method != 'POST':
        return redirect('core:admin_client_detail', pk=client.pk)

    form = ClientPortalDocumentForm(request.POST, request.FILES)
    if form.is_valid():
        document = form.save(commit=False)
        document.client = client
        document.published_by = request.user
        document.save()

        portal_user = _get_client_portal_user(client)
        if portal_user:
            notification_service.create_ui_notification(
                user=portal_user,
                title=f"Nouveau document disponible : {document.title}",
                message="Un nouveau document a été publié dans votre espace client.",
                notification_type='document_updated',
                link_url=reverse('core:client_document_detail', args=[document.pk]),
            )
            # Envoyer un email au client
            if portal_user.email:
                notification_service.send_email_notification(
                    to_emails=[portal_user.email],
                    subject=f"Nouveau document disponible : {document.title}",
                    template_name='notification_generic',
                    context={
                        'headline': 'Nouveau document dans votre espace',
                        'intro': f'Bonjour {portal_user.get_full_name() or portal_user.username},<br><br>Un nouveau document a été publié dans votre espace client : <strong>{document.title}</strong>.',
                        'rows': [
                            {'label': 'Document', 'value': document.title},
                            {'label': 'Type', 'value': document.get_category_display()},
                        ],
                        'action_label': 'Consulter le document',
                    },
                )
            messages.success(request, f"Document publié pour {client.full_name}, notification et email envoyés.")
        else:
            messages.success(request, f"Document publié pour {client.full_name}.")
            messages.warning(request, "Aucun compte portail n'est actif pour ce client. Le document sera visible dès activation de son accès.")
    else:
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)

    return redirect('core:admin_client_detail', pk=client.pk)


@admin_portal_required
def admin_create_invoice(request):
    """Admin Portal invoice creation view."""
    from .forms import InvoiceCreationForm
    prefilled_client = None
    prefilled_quote = None

    if request.GET.get('quote'):
        prefilled_quote = Quote.objects.select_related('client').filter(pk=request.GET.get('quote')).first()
    if request.GET.get('client'):
        prefilled_client = Client.objects.filter(pk=request.GET.get('client')).first()
    elif prefilled_quote:
        prefilled_client = prefilled_quote.client
    
    if request.method == 'POST':
        form = InvoiceCreationForm(request.POST)
        if request.POST.get('quote'):
            prefilled_quote = Quote.objects.select_related('client').filter(pk=request.POST.get('quote')).first()
            if prefilled_quote:
                prefilled_client = prefilled_quote.client

        if prefilled_client:
            form.fields['quote'].queryset = form.fields['quote'].queryset.filter(client=prefilled_client)

        if form.is_valid():
            invoice = form.save()
            messages.success(request, f"Facture {invoice.number} créée avec succès!")
            return redirect('core:admin_invoices_list')
    else:
        initial = {'quote': prefilled_quote.pk} if prefilled_quote else None
        form = InvoiceCreationForm(initial=initial)
        if prefilled_client:
            form.fields['quote'].queryset = form.fields['quote'].queryset.filter(client=prefilled_client)
            if not form.fields['quote'].queryset.exists():
                messages.warning(request, "Aucun devis accepté n'est actuellement disponible pour ce client.")
    
    return render(request, 'core/admin_create_invoice.html', {
        'form': form,
        'prefilled_client': prefilled_client,
        'prefilled_quote': prefilled_quote,
    })


@admin_portal_required
def admin_invoices_list(request):
    """Admin Portal invoices management view with pagination."""
    from django.core.paginator import Paginator
    
    invoices = Invoice.objects.select_related('quote', 'quote__client').order_by('-issue_date')
    paginator = Paginator(invoices, 25)  # 25 invoices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/admin_invoices_list.html', {
        'invoices': page_obj,
        'page_obj': page_obj
    })


@admin_portal_required
def admin_quotes_list(request):
    """Admin Portal quotes management view with pagination."""
    from django.core.paginator import Paginator
    
    quotes = Quote.objects.select_related('client', 'service').order_by('-issue_date')
    paginator = Paginator(quotes, 25)  # 25 quotes per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/admin_quotes_list.html', {
        'quotes': page_obj,
        'page_obj': page_obj
    })


@admin_portal_required
def admin_send_quote_email(request, pk):
    """Admin Portal send quote by email view.
    
    Utilise le template stylisé modele_quote.html pour l'envoi.
    Aucun texte libre n'est permis - le contenu est généré automatiquement.
    """
    from smtplib import SMTPAuthenticationError
    from .forms import QuoteEmailForm
    from django.core.mail import EmailMessage as DjangoEmailMessage
    from django.template.loader import render_to_string
    from django.conf import settings
    from django.urls import reverse
    import logging
    
    logger = logging.getLogger(__name__)
    quote = get_object_or_404(Quote, pk=pk)
    
    if request.method == 'POST':
        form = QuoteEmailForm(quote=quote, data=request.POST)
        if form.is_valid():
            try:
                # Get explicit from_email and branding
                from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@nettoyageexpresse.fr')
                branding = getattr(settings, 'INVOICE_BRANDING', {})
                
                # Générer le sujet automatiquement
                subject = f"Votre devis {quote.number} — {branding.get('name', 'Nettoyage Express')}"
                
                # Générer le contenu HTML avec le template stylisé
                site_url = branding.get('site_url', request.build_absolute_uri('/'))
                html_body = render_to_string('emails/modele_quote.html', {
                    'quote': quote,
                    'client_name': quote.client.full_name,
                    'branding': branding,
                    'pdf_url': request.build_absolute_uri(reverse('devis:download', args=[quote.pk])),
                    'validation_url': request.build_absolute_uri(reverse('devis:quote_validate_start', kwargs={'token': quote.public_token})) if quote.public_token else '#',
                })
                
                # Create email with styled template
                email = DjangoEmailMessage(
                    subject=subject,
                    body=html_body,
                    from_email=from_email,
                    to=[form.cleaned_data['recipient_email']]
                )
                email.content_subtype = 'html'
                
                # Generate PDF fresh (don't rely on stored file - ephemeral filesystem on Render)
                try:
                    pdf_bytes = quote.generate_pdf(attach=False)
                    email.attach(f'devis_{quote.number}.pdf', pdf_bytes, 'application/pdf')
                except Exception as attach_error:
                    logger.error(f"Erreur lors de la génération du PDF pour le devis {quote.number}: {attach_error}")
                    messages.warning(request, f"Le devis a été envoyé mais le PDF n'a pas pu être attaché : {str(attach_error)}")
                
                # Send email with proper error handling
                try:
                    email.send(fail_silently=False)
                except Exception as send_error:
                    logger.error(f"Erreur lors de l'envoi de l'email pour le devis {quote.number}: {send_error}")
                    messages.error(request, f"Erreur lors de l'envoi de l'email : {str(send_error)}")
                    return render(request, 'core/admin_send_quote_email.html', {'form': form, 'quote': quote})
                
                # Update quote status
                if quote.status == Quote.QuoteStatus.DRAFT:
                    quote.status = Quote.QuoteStatus.SENT
                    quote.save(update_fields=['status'])
                
                messages.success(request, f"Devis {quote.number} envoyé avec succès à {form.cleaned_data['recipient_email']}!")
                return redirect('core:admin_quotes_list')
                
            except SMTPAuthenticationError:
                logger.error(f"Erreur auth SMTP devis {quote.number}", exc_info=True)
                messages.error(
                    request,
                    "Authentification SMTP refusée par Brevo. Vérifiez BREVO_SMTP_LOGIN/BREVO_SMTP_PASSWORD ou configurez BREVO_API_KEY.",
                )
            except Exception as e:
                logger.error(f"Erreur inattendue lors de l'envoi de l'email pour le devis {quote.number}: {e}", exc_info=True)
                messages.error(request, f"Erreur lors de l'envoi de l'email : {str(e)}")
    else:
        form = QuoteEmailForm(quote=quote)
    
    return render(request, 'core/admin_send_quote_email.html', {'form': form, 'quote': quote})


@admin_portal_required
def admin_worker_detail(request, pk):
    """Admin Portal worker detail view."""
    from django.contrib.auth.models import User
    from accounts.models import Profile
    from .services.worker_service import WorkerService
    
    worker = get_object_or_404(
        User.objects.select_related('profile'),
        pk=pk,
        profile__role=Profile.ROLE_WORKER
    )
    
    # Récupérer les statistiques du worker
    stats = WorkerService.get_worker_statistics(worker)
    
    # Tâches assignées
    from tasks.models import Task
    tasks = Task.objects.filter(assigned_to=worker).order_by('-due_date', '-created_at')
    
    # Tâches en cours (upcoming + in_progress)
    current_tasks = tasks.filter(status__in=[Task.STATUS_UPCOMING, Task.STATUS_IN_PROGRESS])
    
    # Tâches terminées (dernières 10)
    completed_tasks = tasks.filter(status=Task.STATUS_COMPLETED)[:10]
    
    return render(request, 'core/admin_worker_detail.html', {
        'worker': worker,
        'stats': stats,
        'current_tasks': current_tasks,
        'completed_tasks': completed_tasks,
    })


@admin_portal_required
def admin_client_detail(request, pk):
    """Admin Portal client detail view."""
    from django.urls import reverse
    from .forms import ClientPortalDocumentForm
    from .services.client_service import ClientService
    
    client = get_object_or_404(Client, pk=pk)
    
    # Récupérer les statistiques
    stats = ClientService.get_client_statistics(client)
    
    # Historique
    history = ClientService.get_client_history(client, limit=50)
    
    # Devis associés
    quotes = Quote.objects.filter(client=client).select_related('service').order_by('-issue_date')
    
    # Factures associées (via quote)
    invoices = Invoice.objects.filter(quote__client=client).select_related('quote').order_by('-issue_date')
    
    # Vérifier si un User existe avec cet email
    user_account = _get_client_portal_user(client)
    portal_documents = client.portal_documents.select_related('published_by').all()
    invoice_candidate = quotes.filter(status=Quote.QuoteStatus.ACCEPTED).exclude(invoices__isnull=False).first()
    overdue_invoice = invoices.filter(status=Invoice.InvoiceStatus.OVERDUE).order_by('due_date', '-issue_date').first()
    expiring_quote = quotes.filter(status=Quote.QuoteStatus.SENT, valid_until__isnull=False).order_by('valid_until').first()
    expired_document = portal_documents.filter(expires_at__lt=timezone.localdate()).first()

    client_actions = [
        {
            'label': 'Créer un devis',
            'description': 'Ouvrir un nouveau devis avec le client déjà sélectionné.',
            'url': f"{reverse('core:admin_create_quote')}?client={client.pk}",
            'icon': 'fa-file-circle-plus',
            'tone': 'primary',
        },
        {
            'label': 'Planifier une tâche',
            'description': 'Préremplir une tâche avec ce client, son adresse et son contexte.',
            'url': f"{reverse('core:admin_create_task')}?client={client.pk}",
            'icon': 'fa-calendar-plus',
            'tone': 'slate',
        },
    ]

    if invoice_candidate:
        client_actions.insert(1, {
            'label': f"Facturer {invoice_candidate.number}",
            'description': 'Basculer directement sur une facture prête à partir du devis accepté.',
            'url': f"{reverse('core:admin_create_invoice')}?client={client.pk}&quote={invoice_candidate.pk}",
            'icon': 'fa-receipt',
            'tone': 'emerald',
        })

    if not user_account:
        client_actions.append({
            'label': 'Activer le portail client',
            'description': 'Créer ou relancer l’accès espace client depuis ce dossier.',
            'url': '#portal-access',
            'icon': 'fa-user-lock',
            'tone': 'blue',
        })

    if stats['published_documents_count'] == 0 or stats['expired_documents_count'] > 0:
        client_actions.append({
            'label': 'Mettre à jour les documents',
            'description': 'Publier un document de suivi, rapport ou attestation dans le portail.',
            'url': '#publish-document',
            'icon': 'fa-folder-open',
            'tone': 'amber',
        })

    client_alerts = []

    if overdue_invoice:
        client_alerts.append({
            'title': f"Relance urgente sur {overdue_invoice.number}",
            'description': f"Facture en retard de {overdue_invoice.total_ttc:.2f} € à suivre immédiatement.",
            'url': reverse('core:admin_invoice_detail', args=[overdue_invoice.pk]),
            'cta': 'Ouvrir la facture',
            'icon': 'fa-triangle-exclamation',
            'tone': 'danger',
        })

    if invoice_candidate:
        client_alerts.append({
            'title': f"{invoice_candidate.number} est prêt à être facturé",
            'description': f"Un devis accepté de {invoice_candidate.total_ttc:.2f} € attend sa conversion en facture.",
            'url': f"{reverse('core:admin_create_invoice')}?client={client.pk}&quote={invoice_candidate.pk}",
            'cta': 'Créer la facture',
            'icon': 'fa-circle-check',
            'tone': 'success',
        })

    if expiring_quote and expiring_quote.valid_until and expiring_quote.valid_until <= timezone.localdate() + timedelta(days=7):
        expiry_label = 'a expiré' if expiring_quote.valid_until < timezone.localdate() else 'expire bientôt'
        client_alerts.append({
            'title': f"{expiring_quote.number} {expiry_label}",
            'description': f"Le devis envoyé au client doit être relancé avant le {expiring_quote.valid_until.strftime('%d/%m/%Y')}.",
            'url': reverse('core:admin_quote_detail', args=[expiring_quote.pk]),
            'cta': 'Ouvrir le devis',
            'icon': 'fa-hourglass-half',
            'tone': 'warning',
        })

    if expired_document:
        client_alerts.append({
            'title': 'Document portail à renouveler',
            'description': f"{expired_document.title} est expiré. Une mise à jour évite un portail client obsolète.",
            'url': '#publish-document',
            'cta': 'Publier un document',
            'icon': 'fa-folder-minus',
            'tone': 'warning',
        })

    if stats['days_since_last_activity'] and stats['days_since_last_activity'] >= 45:
        client_alerts.append({
            'title': 'Compte à relancer',
            'description': f"Aucune activité récente depuis {stats['days_since_last_activity']} jours sur ce dossier client.",
            'url': f"{reverse('core:admin_create_quote')}?client={client.pk}",
            'cta': 'Préparer une relance',
            'icon': 'fa-phone-volume',
            'tone': 'slate',
        })
    
    return render(request, 'core/admin_client_detail.html', {
        'client': client,
        'stats': stats,
        'history': history,
        'history_preview': history[:12],
        'quotes': quotes,
        'invoices': invoices,
        'user_account': user_account,
        'portal_documents': portal_documents,
        'portal_document_form': ClientPortalDocumentForm(),
        'client_actions': client_actions,
        'client_alerts': client_alerts,
        'invoice_candidate': invoice_candidate,
    })


@admin_portal_required
def admin_quote_detail(request, pk):
    """Admin Portal quote detail view."""
    quote = get_object_or_404(
        Quote.objects.select_related('client', 'service', 'quote_request'),
        pk=pk
    )
    
    # Lignes du devis
    quote_items = quote.quote_items.all().select_related('service')
    
    # Facture associée (si convertie)
    invoice = quote.invoices.first() if quote.invoices.exists() else None
    
    return render(request, 'core/admin_quote_detail.html', {
        'quote': quote,
        'quote_items': quote_items,
        'invoice': invoice,
    })


@admin_portal_required
def admin_invoice_detail(request, pk):
    """Admin Portal invoice detail view."""
    invoice = get_object_or_404(
        Invoice.objects.select_related('quote', 'quote__client'),
        pk=pk
    )
    
    # Lignes de la facture
    invoice_items = invoice.invoice_items.all()
    
    return render(request, 'core/admin_invoice_detail.html', {
        'invoice': invoice,
        'invoice_items': invoice_items,
    })


@admin_portal_required
def admin_invoice_mark_paid(request, pk):
    """Mark an invoice as paid."""
    from django.contrib import messages
    
    if request.method != 'POST':
        return redirect('core:admin_invoice_detail', pk=pk)
    
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if invoice.status not in [Invoice.InvoiceStatus.PAID]:
        invoice.status = Invoice.InvoiceStatus.PAID
        invoice.save()
        messages.success(request, f"La facture {invoice.number} a été marquée comme payée.")
    else:
        messages.info(request, f"La facture {invoice.number} est déjà payée.")
    
    return redirect('core:admin_invoice_detail', pk=pk)


@admin_portal_required
def admin_edit_invoice(request, pk):
    """Admin Portal invoice edit view."""
    from django.contrib import messages
    from django import forms
    from django.forms import inlineformset_factory
    from factures.models import Invoice, InvoiceItem
    
    invoice = get_object_or_404(Invoice.objects.select_related('quote', 'quote__client'), pk=pk)
    
    # Form pour la facture
    class InvoiceEditForm(forms.ModelForm):
        class Meta:
            model = Invoice
            fields = ['status', 'issue_date', 'due_date', 'discount', 'notes', 'payment_terms']
            widgets = {
                'status': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
                'issue_date': forms.DateInput(attrs={'type': 'date', 'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
                'due_date': forms.DateInput(attrs={'type': 'date', 'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
                'discount': forms.NumberInput(attrs={'step': '0.01', 'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
                'notes': forms.Textarea(attrs={'rows': 3, 'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
                'payment_terms': forms.Textarea(attrs={'rows': 2, 'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:ring-2 focus:ring-ne-primary-500'}),
            }
    
    # Form pour les lignes de facture
    class InvoiceItemForm(forms.ModelForm):
        class Meta:
            model = InvoiceItem
            fields = ['description', 'quantity', 'unit_price', 'tax_rate']
            widgets = {
                'description': forms.TextInput(attrs={'class': 'w-full px-2 py-1.5 text-sm border border-gray-300 rounded focus:outline-none focus:ring-1 focus:ring-ne-primary-500', 'placeholder': 'Description...'}),
                'quantity': forms.NumberInput(attrs={'class': 'w-20 px-2 py-1.5 text-sm border border-gray-300 rounded focus:outline-none focus:ring-1 focus:ring-ne-primary-500 text-right qty-input', 'min': '1', 'step': '1'}),
                'unit_price': forms.NumberInput(attrs={'step': '0.01', 'class': 'w-24 px-2 py-1.5 text-sm border border-gray-300 rounded focus:outline-none focus:ring-1 focus:ring-ne-primary-500 text-right price-input'}),
                'tax_rate': forms.NumberInput(attrs={'step': '0.01', 'value': '0.00', 'class': 'w-20 px-2 py-1.5 text-sm border border-gray-300 rounded focus:outline-none focus:ring-1 focus:ring-ne-primary-500 text-right tva-input'}),
            }
    
    InvoiceItemFormSet = inlineformset_factory(
        Invoice, InvoiceItem,
        form=InvoiceItemForm,
        extra=1,
        can_delete=True
    )
    
    if request.method == 'POST':
        form = InvoiceEditForm(request.POST, instance=invoice)
        formset = InvoiceItemFormSet(request.POST, instance=invoice, prefix='items')
        
        if form.is_valid() and formset.is_valid():
            form.save()
            
            # Handle deletions FIRST - before saving new/updated items
            for item_form in formset:
                if item_form.cleaned_data.get('DELETE') and item_form.instance.pk:
                    item_form.instance.delete()
            
            # Save formset (new and updated items only)
            instances = formset.save(commit=False)
            for instance in instances:
                instance.invoice = invoice
                instance.save()
            
            # Also handle deleted_objects as a fallback
            for obj in formset.deleted_objects:
                if obj.pk:
                    obj.delete()
            
            invoice.compute_totals()
            messages.success(request, f"La facture {invoice.number} a été mise à jour.")
            return redirect('core:admin_invoice_detail', pk=pk)
    else:
        form = InvoiceEditForm(instance=invoice)
        formset = InvoiceItemFormSet(instance=invoice, prefix='items')
    
    return render(request, 'core/admin_edit_invoice.html', {
        'invoice': invoice,
        'form': form,
        'formset': formset,
    })


@admin_portal_required
def admin_send_invoice_email(request, pk):
    """
    Admin Portal - Envoi de facture par email.
    
    GET: Affiche le formulaire de confirmation avec l'email pré-rempli
    POST: Envoie l'email avec le PDF attaché
    """
    from smtplib import SMTPAuthenticationError
    from django.core.mail import EmailMessage as DjangoEmailMessage
    from django.template.loader import render_to_string
    import logging
    
    logger = logging.getLogger(__name__)
    invoice = get_object_or_404(
        Invoice.objects.select_related('quote', 'quote__client'), 
        pk=pk
    )
    
    # Récupérer les infos client
    client_email = ''
    client_name = 'Client'
    if invoice.quote and invoice.quote.client:
        client_email = invoice.quote.client.email or ''
        client_name = invoice.quote.client.full_name
    
    if request.method == 'POST':
        recipient_email = request.POST.get('recipient_email', '').strip()
        
        if not recipient_email:
            messages.error(request, "Veuillez spécifier une adresse email.")
            return render(request, 'core/admin_send_invoice_email.html', {
                'invoice': invoice,
                'client_email': client_email,
                'client_name': client_name,
            })
        
        try:
            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@nettoyageexpresse.fr')
            branding = getattr(settings, 'INVOICE_BRANDING', {})
            company_name = branding.get('name', 'Nettoyage Express')
            
            # Générer le contenu HTML
            html_body = render_to_string('emails/invoice_notification.html', {
                'invoice': invoice,
                'client_name': client_name,
                'branding': branding,
            })
            
            subject = f"Votre facture {invoice.number} — {company_name}"
            
            # Créer l'email
            email = DjangoEmailMessage(
                subject=subject,
                body=html_body,
                from_email=from_email,
                to=[recipient_email]
            )
            email.content_subtype = 'html'
            
            # Attacher le PDF
            try:
                pdf_bytes = invoice.generate_pdf(attach=False)
                email.attach(f'facture_{invoice.number}.pdf', pdf_bytes, 'application/pdf')
            except Exception as pdf_error:
                logger.error(f"Erreur génération PDF facture {invoice.number}: {pdf_error}")
                messages.warning(request, "Le PDF n'a pas pu être généré, l'email sera envoyé sans pièce jointe.")
            
            # Envoyer
            email.send(fail_silently=False)
            
            # Mettre à jour le statut si brouillon
            if invoice.status == Invoice.InvoiceStatus.DRAFT:
                invoice.status = Invoice.InvoiceStatus.SENT
                invoice.save(update_fields=['status'])
            
            messages.success(request, f"Facture {invoice.number} envoyée avec succès à {recipient_email}!")
            return redirect('core:admin_invoice_detail', pk=pk)
            
        except SMTPAuthenticationError:
            logger.error(f"Erreur auth SMTP facture {invoice.number}", exc_info=True)
            messages.error(
                request,
                "Authentification SMTP refusée par Brevo. Vérifiez BREVO_SMTP_LOGIN/BREVO_SMTP_PASSWORD ou configurez BREVO_API_KEY.",
            )
            return render(request, 'core/admin_send_invoice_email.html', {
                'invoice': invoice,
                'client_email': client_email,
                'client_name': client_name,
            })
        except Exception as e:
            logger.error(f"Erreur envoi email facture {invoice.number}: {e}", exc_info=True)
            messages.error(request, f"Erreur lors de l'envoi: {str(e)}")
            return render(request, 'core/admin_send_invoice_email.html', {
                'invoice': invoice,
                'client_email': client_email,
                'client_name': client_name,
            })
    
    # GET - Afficher le formulaire
    return render(request, 'core/admin_send_invoice_email.html', {
        'invoice': invoice,
        'client_email': client_email,
        'client_name': client_name,
    })


@admin_portal_required
def admin_task_detail(request, pk):
    """Admin Portal task detail view."""
    from tasks.models import Task
    
    task = get_object_or_404(
        Task.objects.prefetch_related('assigned_to').select_related('client', 'completed_by'),
        pk=pk
    )
    
    return render(request, 'core/admin_task_detail.html', {
        'task': task,
    })


@admin_portal_required
def admin_tasks_list(request):
    """Admin Portal tasks list view."""
    from tasks.models import Task
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    tasks = Task.objects.prefetch_related('assigned_to').select_related('client').order_by('-due_date', '-created_at')
    
    # Filtres
    status_filter = request.GET.get('status')
    worker_filter = request.GET.get('worker')
    client_filter = request.GET.get('client')
    search_query = request.GET.get('q')  # 'q' pour correspondre au template
    
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    
    if worker_filter:
        if worker_filter == 'unassigned':
            # Tâches non assignées (pas d'ouvriers)
            from django.db.models import Count
            tasks = tasks.annotate(worker_count=Count('assigned_to')).filter(worker_count=0)
        else:
            tasks = tasks.filter(assigned_to__id=worker_filter)

    if client_filter:
        tasks = tasks.filter(client__id=client_filter)
    
    if search_query:
        tasks = tasks.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(client__full_name__icontains=search_query) |
            Q(client__company__icontains=search_query) |
            Q(client__email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(tasks, 25)  # 25 tasks per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Liste des workers pour le filtre
    from django.contrib.auth.models import User
    from accounts.models import Profile
    workers = User.objects.filter(profile__role=Profile.ROLE_WORKER, is_active=True).order_by('first_name', 'last_name')
    clients = Client.objects.order_by('company', 'full_name')
    
    return render(request, 'core/admin_tasks_list.html', {
        'tasks': page_obj,
        'page_obj': page_obj,
        'workers': workers,
        'clients': clients,
        'status_filter': status_filter,
        'worker_filter': worker_filter,
        'client_filter': client_filter,
        'search_query': search_query,
    })


@admin_portal_required
def admin_convert_quote_to_invoice(request, pk):
    """Admin Portal convert quote to invoice view."""
    from devis.services import create_invoice_from_quote, QuoteAlreadyInvoicedError, QuoteStatusError
    from django.http import Http404
    
    quote = get_object_or_404(Quote, pk=pk)
    
    if request.method == 'POST':
        try:
            result = create_invoice_from_quote(quote)
            messages.success(request, f"Facture {result.invoice.number} créée avec succès depuis le devis {quote.number}!")
            return redirect('core:admin_invoice_detail', pk=result.invoice.pk)
        except QuoteAlreadyInvoicedError as e:
            messages.error(request, f"Ce devis a déjà été facturé.")
            return redirect('core:admin_quote_detail', pk=pk)
        except QuoteStatusError as e:
            messages.error(request, f"Ce devis ne peut pas être converti : {str(e)}")
            return redirect('core:admin_quote_detail', pk=pk)
    
    # GET: Afficher confirmation
    # Vérifier si déjà facturé
    if quote.invoices.exists():
        messages.warning(request, "Ce devis a déjà été facturé.")
        return redirect('core:admin_quote_detail', pk=pk)
    
    # Vérifier statut
    if quote.status != Quote.QuoteStatus.ACCEPTED:
        messages.error(request, f"Seuls les devis acceptés peuvent être convertis en facture. Statut actuel : {quote.get_status_display()}")
        return redirect('core:admin_quote_detail', pk=pk)
    
    return render(request, 'core/admin_convert_quote_to_invoice.html', {
        'quote': quote,
    })


# Notification HTMX Views

@login_required
def notification_count(request):
    """HTMX endpoint to get unread notification count."""
    from core.models import UINotification
    
    count = UINotification.get_unread_count(request.user)
    return render(
        request,
        "core/partials/notification_count.html",
        {"count": count}
    )


@login_required
def notification_list(request):
    """HTMX endpoint to get notification list."""
    from core.models import UINotification
    
    notifications = UINotification.get_recent_notifications(request.user, limit=10)
    return render(
        request,
        "core/partials/notification_list.html",
        {"notifications": notifications}
    )


@login_required
def mark_notification_read(request, notification_id):
    """HTMX endpoint to mark a notification as read."""
    from core.models import UINotification
    
    try:
        notification = UINotification.objects.get(id=notification_id, user=request.user)
        notification.mark_as_read()
    except UINotification.DoesNotExist:
        pass
    
    # Return updated notification list for HTMX swap
    notifications = UINotification.get_recent_notifications(request.user, limit=10)
    return render(
        request,
        "core/partials/notification_list.html",
        {"notifications": notifications}
    )


@login_required
def mark_all_notifications_read(request):
    """HTMX endpoint to mark all notifications as read."""
    from core.models import UINotification
    from django.utils import timezone
    
    UINotification.objects.filter(user=request.user, read_at__isnull=True).update(
        read_at=timezone.now()
    )
    
    # Return updated notification list
    notifications = UINotification.get_recent_notifications(request.user, limit=10)
    return render(
        request,
        "core/partials/notification_list.html",
        {"notifications": notifications}
    )


# ============================================================================
# ANALYTICS & REPORTING VIEWS
# ============================================================================

@login_required
def admin_analytics(request):
    """
    Tableau de bord analytique avancé avec KPIs détaillés et graphiques.
    """
    from accounts.portal import get_user_role
    from core.services.analytics_service import AnalyticsService
    
    role = get_user_role(request.user)
    if role not in ['admin_technical', 'admin_business']:
        return redirect('core:access_denied')
    
    # Période sélectionnée (défaut: mois)
    period = request.GET.get('period', 'month')
    
    # KPIs avancés avec comparaison
    kpis = AnalyticsService.get_advanced_kpis(period)
    
    # Comparaison mensuelle année courante vs précédente
    monthly_comparison = AnalyticsService.get_monthly_comparison()
    
    # Funnel de conversion devis
    quote_funnel = AnalyticsService.get_quote_funnel()
    
    # Top services
    top_services = AnalyticsService.get_services_demand()
    
    # Top clients
    top_clients = AnalyticsService.get_revenue_by_client(limit=10)
    
    # Performance workers
    worker_stats = AnalyticsService.get_worker_detailed_stats()
    
    # Factures en retard
    overdue_invoices = AnalyticsService.get_overdue_invoices()[:10]
    
    context = {
        'period': period,
        'kpis': kpis,
        'monthly_comparison': monthly_comparison,
        'quote_funnel': quote_funnel,
        'top_services': top_services,
        'top_clients': top_clients,
        'worker_stats': worker_stats,
        'overdue_invoices': overdue_invoices,
    }
    
    return render(request, 'core/admin_analytics.html', context)


@login_required
def admin_reports(request):
    """
    Page de génération de rapports personnalisés.
    """
    from accounts.portal import get_user_role
    from core.services.analytics_service import ReportingService
    from datetime import datetime, timedelta
    
    role = get_user_role(request.user)
    if role not in ['admin_technical', 'admin_business']:
        return redirect('core:access_denied')
    
    # Dates par défaut (mois en cours)
    today = timezone.now().date()
    default_start = today.replace(day=1)
    default_end = today
    
    report_type = request.GET.get('type', None)
    start_date = request.GET.get('start_date', default_start.isoformat())
    end_date = request.GET.get('end_date', default_end.isoformat())
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start
        end_date = default_end
    
    report_data = None
    
    if report_type == 'revenue':
        report_data = ReportingService.generate_revenue_report(start_date, end_date)
    elif report_type == 'clients':
        report_data = ReportingService.generate_client_report(start_date, end_date)
    elif report_type == 'workers':
        report_data = ReportingService.generate_worker_report(start_date, end_date)
    
    context = {
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date,
        'report_data': report_data,
    }
    
    return render(request, 'core/admin_reports.html', context)


@login_required
def admin_export_report(request):
    """
    Export de rapport en CSV.
    """
    from accounts.portal import get_user_role
    from core.services.analytics_service import ReportingService
    from django.http import HttpResponse
    from datetime import datetime
    import csv
    
    role = get_user_role(request.user)
    if role not in ['admin_technical', 'admin_business']:
        return redirect('core:access_denied')
    
    report_type = request.GET.get('type', 'revenue')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="rapport_{report_type}_{start_date}_{end_date}.csv"'
    response.write('\ufeff')  # BOM for Excel UTF-8
    
    writer = csv.writer(response, delimiter=';')
    
    if report_type == 'revenue':
        report = ReportingService.generate_revenue_report(start_date, end_date)
        writer.writerow(['Numero', 'Date', 'Client', 'Montant HT', 'TVA', 'Montant TTC', 'Statut'])
        for inv in report['invoices']:
            client_name = inv.quote.client.full_name if inv.quote and inv.quote.client else 'N/A'
            writer.writerow([
                inv.number,
                inv.issue_date.strftime('%d/%m/%Y'),
                client_name,
                str(inv.total_ht).replace('.', ','),
                str(inv.tva).replace('.', ','),
                str(inv.total_ttc).replace('.', ','),
                inv.get_status_display()
            ])
        writer.writerow([])
        writer.writerow(['TOTAL', '', '', str(report['total_ht']).replace('.', ','), 
                        str(report['total_tva']).replace('.', ','), str(report['total_ttc']).replace('.', ',')])
    
    elif report_type == 'clients':
        report = ReportingService.generate_client_report(start_date, end_date)
        writer.writerow(['Client', 'Entreprise', 'Nombre Devis', 'CA Total'])
        for client in report['clients']:
            writer.writerow([
                client.full_name,
                client.company or '',
                client.quote_count,
                str(client.invoice_total or 0).replace('.', ',')
            ])
    
    elif report_type == 'workers':
        report = ReportingService.generate_worker_report(start_date, end_date)
        writer.writerow(['Ouvrier', 'Taches Totales', 'Terminees', 'En Retard', 'Taux Completion'])
        for worker in report['workers']:
            completion = (worker.report_completed_tasks / worker.report_total_tasks * 100) if worker.report_total_tasks > 0 else 0
            writer.writerow([
                worker.get_full_name() or worker.username,
                worker.report_total_tasks,
                worker.report_completed_tasks,
                worker.report_overdue_tasks,
                f"{completion:.1f}%"
            ])
    
    return response
